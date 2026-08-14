"""Leitura e importação do mapa de agendamentos em formato Excel."""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.extensions import db
from app.models import AgendamentoLaboratorio, Laboratorio, Usuario


TURNOS_EXCEL = {
    "manha": (time(8, 0), time(12, 0)),
    "tarde": (time(13, 0), time(17, 0)),
    "noite": (time(18, 0), time(22, 0)),
}
PLACEHOLDERS = {"0", "/", "-", "—"}


class AgendaExcelError(ValueError):
    """Erro de validação que pode ser exibido diretamente no terminal."""


@dataclass(frozen=True)
class AgendaExcelEntry:
    data: date
    turno: str
    laboratorio_nome: str
    texto: str
    aba: str
    celula: str


@dataclass
class AgendaExcelData:
    registros: list[AgendaExcelEntry] = field(default_factory=list)
    placeholders_ignorados: int = 0
    avisos: list[str] = field(default_factory=list)


@dataclass
class AgendaImportResult:
    encontrados: int
    prontos: int
    importados: int
    conflitos: int
    duplicados_na_planilha: int
    placeholders_ignorados: int
    avisos: list[str]


def _normalize(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_date(value, epoch) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        parsed = from_excel(value, epoch)
        return parsed.date() if isinstance(parsed, datetime) else parsed

    raw = _clean_text(value)
    for date_format in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    raise AgendaExcelError(f"data não reconhecida: {raw or value!r}")


def _find_columns(worksheet):
    max_scan_row = min(worksheet.max_row, 20)
    for row_number in range(1, max_scan_row + 1):
        headings = {
            column: _normalize(worksheet.cell(row_number, column).value)
            for column in range(1, worksheet.max_column + 1)
        }
        data_column = next(
            (column for column, heading in headings.items() if heading == "data"), None
        )
        shift_column = next(
            (column for column, heading in headings.items() if heading == "turno"), None
        )
        lab_columns = {}
        for column, heading in headings.items():
            match = re.fullmatch(r"lab(?:oratorio)?\s*0?([1-5])", heading)
            if match:
                lab_columns[column] = f"Lab {int(match.group(1)):02d}"
        if data_column and shift_column and lab_columns:
            return row_number, data_column, shift_column, lab_columns
    return None


def read_agenda_workbook(path: str | Path) -> AgendaExcelData:
    """Extrai reservas de todas as abas sem modificar a planilha."""
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise AgendaExcelError(f"Arquivo não encontrado: {workbook_path}")

    workbook = load_workbook(
        filename=workbook_path,
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    result = AgendaExcelData()
    try:
        for worksheet in workbook.worksheets:
            columns = _find_columns(worksheet)
            if not columns:
                result.avisos.append(
                    f"Aba {worksheet.title!r} ignorada: cabeçalho Data/Turno/Lab não encontrado."
                )
                continue

            header_row, data_column, shift_column, lab_columns = columns
            current_date = None
            for row_number in range(header_row + 1, worksheet.max_row + 1):
                date_value = worksheet.cell(row_number, data_column).value
                if date_value not in (None, ""):
                    try:
                        current_date = _parse_date(date_value, workbook.epoch)
                    except AgendaExcelError as error:
                        current_date = None
                        result.avisos.append(
                            f"{worksheet.title}!{worksheet.cell(row_number, data_column).coordinate}: {error}"
                        )

                shift = _normalize(worksheet.cell(row_number, shift_column).value)
                if shift not in TURNOS_EXCEL:
                    continue
                if current_date is None:
                    result.avisos.append(
                        f"{worksheet.title}!{worksheet.cell(row_number, shift_column).coordinate}: "
                        "turno sem data correspondente."
                    )
                    continue

                for column, lab_name in lab_columns.items():
                    cell = worksheet.cell(row_number, column)
                    text = _clean_text(cell.value)
                    if not text:
                        continue
                    if text in PLACEHOLDERS:
                        result.placeholders_ignorados += 1
                        continue
                    result.registros.append(
                        AgendaExcelEntry(
                            data=current_date,
                            turno=shift,
                            laboratorio_nome=lab_name,
                            texto=text,
                            aba=worksheet.title,
                            celula=cell.coordinate,
                        )
                    )
    finally:
        workbook.close()
    return result


def _resolve_user(email: str | None) -> Usuario:
    query = Usuario.query.filter_by(ativo=True)
    if email:
        user = query.filter(db.func.lower(Usuario.email) == email.strip().lower()).first()
        if not user:
            raise AgendaExcelError(f"Usuário ativo não encontrado: {email}")
        return user

    user = query.filter(Usuario.perfil.in_(("admin", "tecnico"))).order_by(Usuario.id).first()
    if not user:
        raise AgendaExcelError(
            "Nenhum administrador ou técnico ativo foi encontrado para ser o responsável."
        )
    return user


def import_agenda_workbook(
    path: str | Path,
    *,
    user_email: str | None = None,
    commit: bool = False,
) -> AgendaImportResult:
    """Simula ou grava a planilha, ignorando reservas que conflitam com o banco."""
    workbook_path = Path(path)
    spreadsheet = read_agenda_workbook(workbook_path)
    user = _resolve_user(user_email)

    lab_names = sorted({entry.laboratorio_nome for entry in spreadsheet.registros})
    laboratories = Laboratorio.query.filter(Laboratorio.nome.in_(lab_names)).all()
    laboratories_by_name = {laboratory.nome: laboratory for laboratory in laboratories}
    missing_labs = sorted(set(lab_names) - set(laboratories_by_name))
    if missing_labs:
        raise AgendaExcelError(
            "Laboratórios não cadastrados: "
            + ", ".join(missing_labs)
            + ". Execute python init_db.py antes da importação."
        )

    conflicts_by_day = defaultdict(list)
    if spreadsheet.registros:
        first_day = min(entry.data for entry in spreadsheet.registros)
        last_day = max(entry.data for entry in spreadsheet.registros)
        existing = AgendamentoLaboratorio.query.filter(
            AgendamentoLaboratorio.status == "agendado",
            AgendamentoLaboratorio.inicio >= datetime.combine(first_day, time.min),
            AgendamentoLaboratorio.inicio
            < datetime.combine(last_day + timedelta(days=1), time.min),
        ).all()
        for booking in existing:
            conflicts_by_day[(booking.laboratorio_id, booking.inicio.date())].append(
                (booking.inicio, booking.fim)
            )

    pending = []
    source_keys = set()
    conflicts = 0
    duplicates = 0
    for entry in spreadsheet.registros:
        laboratory = laboratories_by_name[entry.laboratorio_nome]
        start_time, end_time = TURNOS_EXCEL[entry.turno]
        start = datetime.combine(entry.data, start_time)
        end = datetime.combine(entry.data, end_time)
        source_key = (laboratory.id, start, end)
        if source_key in source_keys:
            duplicates += 1
            continue
        source_keys.add(source_key)

        has_conflict = any(
            existing_start < end and existing_end > start
            for existing_start, existing_end in conflicts_by_day[(laboratory.id, entry.data)]
        )
        if has_conflict:
            conflicts += 1
            continue

        source_note = (
            f"Importado de {workbook_path.name} · aba {entry.aba} · célula {entry.celula}."
        )
        if len(entry.texto) > 150:
            source_note += f" Texto original: {entry.texto}"
        pending.append(
            AgendamentoLaboratorio(
                laboratorio_id=laboratory.id,
                usuario_id=user.id,
                titulo=entry.texto[:150],
                finalidade=source_note,
                inicio=start,
                fim=end,
                status="agendado",
            )
        )
        conflicts_by_day[(laboratory.id, entry.data)].append((start, end))

    imported = 0
    if commit and pending:
        try:
            db.session.add_all(pending)
            db.session.commit()
            imported = len(pending)
        except Exception:
            db.session.rollback()
            raise

    return AgendaImportResult(
        encontrados=len(spreadsheet.registros),
        prontos=len(pending),
        importados=imported,
        conflitos=conflicts,
        duplicados_na_planilha=duplicates,
        placeholders_ignorados=spreadsheet.placeholders_ignorados,
        avisos=spreadsheet.avisos,
    )
